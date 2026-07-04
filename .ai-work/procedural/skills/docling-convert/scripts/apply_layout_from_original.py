"""
apply_layout_from_original.py
Copy column widths, row heights, page setup, and default dimensions
from the original xlsx into the reconstructed final xlsx.

This is the AI post-processing step: the reconstructed file has correct
cell data and styles, but wrong row heights / column widths because
the markdown pipeline strips empty rows and misses default-width columns.
This script fixes that by reading exact values from the original.

Usage:
    python apply_layout_from_original.py <original.xlsx> <final.xlsx> [--out patched.xlsx]
"""
import argparse
import sys
from pathlib import Path


def copy_layout(original_path: Path, final_path: Path, output_path: Path):
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl not installed.", file=sys.stderr)
        sys.exit(1)

    print(f"Original: {original_path.name}")
    print(f"Final:    {final_path.name}")
    print(f"Output:   {output_path.name}\n")

    wb_orig = openpyxl.load_workbook(str(original_path), data_only=True)
    wb_final = openpyxl.load_workbook(str(final_path))

    for sheet_name in wb_final.sheetnames:
        ws_final = wb_final[sheet_name]
        if sheet_name not in wb_orig.sheetnames:
            print(f"  [{sheet_name}] not in original — skipped")
            continue
        ws_orig = wb_orig[sheet_name]

        changes = []

        # ── Default column width ──────────────────────────────────────────
        orig_default_col = (ws_orig.sheet_format.defaultColWidth
                            or ws_orig.sheet_format.baseColWidth
                            or 8.43)
        if orig_default_col:
            ws_final.sheet_format.defaultColWidth = orig_default_col
            ws_final.sheet_format.baseColWidth = int(orig_default_col)

        # ── Column widths ─────────────────────────────────────────────────
        max_col = ws_orig.max_column or 0
        col_count = 0
        for col_idx in range(1, (max_col or 0) + 1):
            letter = get_column_letter(col_idx)
            orig_dim = ws_orig.column_dimensions.get(letter)
            if orig_dim and orig_dim.width and orig_dim.customWidth:
                ws_final.column_dimensions[letter].width = orig_dim.width
                col_count += 1
            elif orig_dim and orig_dim.hidden:
                ws_final.column_dimensions[letter].hidden = True
        if col_count:
            changes.append(f"{col_count} col widths")

        # ── Default row height ────────────────────────────────────────────
        orig_default_row = ws_orig.sheet_format.defaultRowHeight
        if orig_default_row:
            ws_final.sheet_format.defaultRowHeight = orig_default_row
            ws_final.sheet_format.customHeight = ws_orig.sheet_format.customHeight

        # ── Row heights ───────────────────────────────────────────────────
        row_count = 0
        for row_idx, orig_rd in ws_orig.row_dimensions.items():
            if orig_rd.height is not None:
                ws_final.row_dimensions[row_idx].height = orig_rd.height
                row_count += 1
            if orig_rd.hidden:
                ws_final.row_dimensions[row_idx].hidden = True
        if row_count:
            changes.append(f"{row_count} row heights")

        # ── Page setup ────────────────────────────────────────────────────
        ops = ws_orig.page_setup
        fps = ws_final.page_setup
        for attr in ("orientation", "paperSize", "scale", "fitToPage",
                     "fitToWidth", "fitToHeight"):
            val = getattr(ops, attr, None)
            if val is not None:
                setattr(fps, attr, val)

        opm = ws_orig.page_margins
        fpm = ws_final.page_margins
        for attr in ("left", "right", "top", "bottom", "header", "footer"):
            val = getattr(opm, attr, None)
            if val is not None:
                setattr(fpm, attr, val)

        # Print area
        if ws_orig.print_area:
            ws_final.print_area = ws_orig.print_area

        # Freeze panes
        if ws_orig.freeze_panes:
            ws_final.freeze_panes = ws_orig.freeze_panes

        changes.append("page setup")
        print(f"  [{sheet_name}]: {', '.join(changes)}")

    wb_orig.close()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb_final.save(str(output_path))
    print(f"\nSaved: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Copy column widths, row heights, page setup from original to reconstructed xlsx"
    )
    parser.add_argument("original", help="Original .xlsx source")
    parser.add_argument("final",    help="Reconstructed final .xlsx to patch")
    parser.add_argument("--out",    help="Output path (default: <final>_patched.xlsx)")
    args = parser.parse_args()

    original_path = Path(args.original)
    final_path    = Path(args.final)
    if not original_path.exists():
        print(f"ERROR: not found: {original_path}", file=sys.stderr); sys.exit(1)
    if not final_path.exists():
        print(f"ERROR: not found: {final_path}", file=sys.stderr); sys.exit(1)

    out = Path(args.out) if args.out else final_path.with_name(
        final_path.stem + "_patched" + final_path.suffix
    )
    copy_layout(original_path, final_path, out)


if __name__ == "__main__":
    main()
