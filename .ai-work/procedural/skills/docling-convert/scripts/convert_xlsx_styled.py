"""
convert_xlsx_styled.py
Convert Excel (xlsx/xlsm) → Markdown. Two modes, chosen once and saved.

  Mode 1 - extract   : Plain MD tables only. Fast, LLM-readable, no styling.
  Mode 2 - roundtrip : MD tables + styles.json style guide + screenshot requests.
                       Use apply_style_guide.py to reconstruct styled Excel.

Mode is saved to ~/.docling_convert_config.json. Pass --reset to change.

Output:
  extract    → <stem>_data.md
  roundtrip  → <stem>_data.md  +  <stem>_styles.json

Requires:
    pip install openpyxl

Usage:
    python convert_xlsx_styled.py input.xlsx [--out OUTPUT] [--reset]
"""
import argparse
import json
import re
import sys
import xml.etree.ElementTree as ET
from collections import Counter
from pathlib import Path


# ── Indexed color table (legacy Excel palette) ────────────────────────────────

_INDEXED_COLORS = {
    0: "000000", 1: "FFFFFF", 2: "FF0000", 3: "00FF00", 4: "0000FF",
    5: "FFFF00", 6: "FF00FF", 7: "00FFFF", 8: "000000", 9: "FFFFFF",
    10: "FF0000", 11: "00FF00", 12: "0000FF", 13: "FFFF00", 14: "FF00FF",
    15: "00FFFF", 16: "800000", 17: "008000", 18: "000080", 19: "808000",
    20: "800080", 21: "008080", 22: "C0C0C0", 23: "808080", 24: "9999FF",
    25: "993366", 26: "FFFFCC", 27: "CCFFFF", 28: "660066", 29: "FF8080",
    30: "0066CC", 31: "CCCCFF", 32: "000080", 33: "FF00FF", 34: "FFFF00",
    35: "00FFFF", 36: "800080", 37: "800000", 38: "008080", 39: "0000FF",
    40: "00CCFF", 41: "CCFFFF", 42: "CCFFCC", 43: "FFFF99", 44: "99CCFF",
    45: "FF99CC", 46: "CC99FF", 47: "FFCC99", 48: "3366FF", 49: "33CCCC",
    50: "99CC00", 51: "FFCC00", 52: "FF9900", 53: "FF6600", 54: "666699",
    55: "969696", 56: "003366", 57: "339966", 58: "003300", 59: "333300",
    60: "993300", 61: "993366", 62: "333399", 63: "333333", 64: "000000",
}

# OOXML cell-format theme indices (NOT the clrScheme child order).
# Cell theme="0" = lt1 (background), theme="1" = dk1 (text),
# theme="2" = lt2, theme="3" = dk2.  Accents 4-9 are unchanged.
# This is the opposite of the clrScheme XML child order (dk1, lt1, ...).
_THEME_SLOTS = [
    "lt1", "dk1", "lt2", "dk2",
    "accent1", "accent2", "accent3", "accent4", "accent5", "accent6",
    "hlink", "folHlink",
]


# ── Theme color resolution ────────────────────────────────────────────────────

def _resolve_theme_colors(wb) -> dict[int, str]:
    """Parse workbook theme XML → {slot_index: hex_color}."""
    if not hasattr(wb, "loaded_theme") or not wb.loaded_theme:
        return {}
    try:
        ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
        root = ET.fromstring(wb.loaded_theme)
        scheme = root.find(".//a:clrScheme", ns)
        if scheme is None:
            return {}
        colors: dict[int, str] = {}
        for i, slot in enumerate(_THEME_SLOTS):
            el = scheme.find(f"a:{slot}", ns)
            if el is None:
                continue
            srgb = el.find("a:srgbClr", ns)
            if srgb is not None:
                colors[i] = srgb.get("val", "").upper()
                continue
            sys_clr = el.find("a:sysClr", ns)
            if sys_clr is not None:
                colors[i] = sys_clr.get("lastClr", "").upper()
        return colors
    except Exception:
        return {}


def _apply_tint(hex6: str, tint: float) -> str:
    """Apply Excel luminance tint to a 6-char hex color."""
    r = int(hex6[0:2], 16)
    g = int(hex6[2:4], 16)
    b = int(hex6[4:6], 16)
    if tint > 0:
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)
    elif tint < 0:
        r = int(r * (1 + tint))
        g = int(g * (1 + tint))
        b = int(b * (1 + tint))
    return f"{min(255,r):02X}{min(255,g):02X}{min(255,b):02X}"


# ── Color helpers ─────────────────────────────────────────────────────────────

def _resolve_color(color_obj, theme_colors: dict | None = None, for_font: bool = False) -> str | None:
    """Return 6-char uppercase hex or None.

    for_font=True: also skip explicit black (FF000000) since that is the default text color.
    Theme dk1/lt1 at tint=0 are always skipped - openpyxl default placeholders, not real choices.
    """
    if color_obj is None:
        return None
    try:
        t = color_obj.type
        if t == "rgb":
            rgb = color_obj.rgb  # ARGB e.g. 'FF4472C4'
            if not rgb or len(rgb) != 8:
                return None
            if rgb[:2] == "00":
                return None  # transparent = default/unset
            if rgb == "FFFFFFFF":
                return None  # explicit white - default background, skip
            if for_font and rgb == "FF000000":
                return None  # default black text - not an explicit font color choice
            return rgb[-6:].upper()
        elif t == "indexed":
            idx = color_obj.indexed
            if idx is not None and idx != 64:
                return _INDEXED_COLORS.get(idx)
        elif t == "theme":
            if theme_colors:
                idx = color_obj.theme
                tint = color_obj.tint or 0.0
                # dk1 (0) and lt1 (1) at tint=0: openpyxl default placeholders for both fill and font
                if idx in (0, 1) and tint == 0.0:
                    return None
                base = theme_colors.get(idx)
                if base:
                    return _apply_tint(base, tint) if tint else base
    except Exception:
        pass
    return None


def _is_white_or_none(hex_color: str | None) -> bool:
    if hex_color is None:
        return True
    return hex_color.upper() in ("FFFFFF", "000000FF", "")


# ── Border helpers ────────────────────────────────────────────────────────────

def _border_css(side, theme_colors=None) -> str:
    if side is None or side.border_style is None:
        return "none"
    s = side.border_style
    color = _resolve_color(side.color, theme_colors) if side.color else None
    c = f"#{color}" if color else "#999"
    if s in ("thin", "hair", "dashDot", "dashDotDot", "dashed", "dotted"):
        return f"1px solid {c}"
    if s in ("medium", "mediumDashDot", "mediumDashDotDot", "mediumDashed"):
        return f"2px solid {c}"
    if s == "thick":
        return f"3px solid {c}"
    if s == "double":
        return f"3px double {c}"
    return f"1px solid {c}"


def _border_style_name(side) -> str | None:
    """Normalize border style to thin/medium/thick/double or None."""
    if side is None or side.border_style is None:
        return None
    s = side.border_style
    if s in ("thin", "hair", "dashDot", "dashDotDot", "dashed", "dotted"):
        return "thin"
    if s in ("medium", "mediumDashDot", "mediumDashDotDot", "mediumDashed"):
        return "medium"
    if s == "thick":
        return "thick"
    if s == "double":
        return "double"
    return "thin"


# ── Merge range helpers ───────────────────────────────────────────────────────

def _build_merge_map(ws) -> tuple[dict, set]:
    merge_spans: dict[tuple[int, int], tuple[int, int]] = {}
    merged_skip: set[tuple[int, int]] = set()
    for rng in ws.merged_cells.ranges:
        r1, c1, r2, c2 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        merge_spans[(r1, c1)] = (r2 - r1 + 1, c2 - c1 + 1)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if (r, c) != (r1, c1):
                    merged_skip.add((r, c))
    return merge_spans, merged_skip


# ── Cell style → CSS (for full mode) ─────────────────────────────────────────

def _cell_css(cell, no_borders: bool = False, theme_colors=None) -> str:
    styles = []

    fill = cell.fill
    if fill and fill.fill_type not in (None, "none"):
        bg = _resolve_color(fill.fgColor, theme_colors)
        if bg and not _is_white_or_none(bg):
            styles.append(f"background-color:#{bg}")

    font = cell.font
    if font:
        if font.bold:
            styles.append("font-weight:bold")
        if font.italic:
            styles.append("font-style:italic")
        if font.size and font.size != 11:
            styles.append(f"font-size:{int(font.size)}pt")
        fc = _resolve_color(font.color, theme_colors, for_font=True)
        if fc:
            styles.append(f"color:#{fc}")

    align = cell.alignment
    if align and align.horizontal in ("center", "right"):
        styles.append(f"text-align:{align.horizontal}")
    if align and align.wrap_text:
        styles.append("white-space:pre-wrap")

    if not no_borders and cell.border:
        b = cell.border
        top    = _border_css(b.top, theme_colors)
        bottom = _border_css(b.bottom, theme_colors)
        left   = _border_css(b.left, theme_colors)
        right  = _border_css(b.right, theme_colors)
        if any(x != "none" for x in (top, bottom, left, right)):
            if top == bottom == left == right:
                styles.append(f"border:{top}")
            else:
                if top    != "none": styles.append(f"border-top:{top}")
                if bottom != "none": styles.append(f"border-bottom:{bottom}")
                if left   != "none": styles.append(f"border-left:{left}")
                if right  != "none": styles.append(f"border-right:{right}")

    return "; ".join(styles)


# ── Cell style → dict (for guide mode) ───────────────────────────────────────

def _extract_style_dict(cell, theme_colors=None) -> dict:
    """Extract all style properties as a plain dict."""
    sd: dict = {
        "bg_color": None,
        "bold": False,
        "italic": False,
        "font_size": None,
        "font_color": None,
        "h_align": None,
        "wrap_text": False,
        "border_top": None,
        "border_bottom": None,
        "border_left": None,
        "border_right": None,
    }

    fill = cell.fill
    if fill and fill.fill_type not in (None, "none"):
        bg = _resolve_color(fill.fgColor, theme_colors)
        if bg and not _is_white_or_none(bg):
            sd["bg_color"] = bg

    font = cell.font
    if font:
        sd["bold"] = bool(font.bold)
        sd["italic"] = bool(font.italic)
        if font.size and font.size != 11:
            sd["font_size"] = int(font.size)
        fc = _resolve_color(font.color, theme_colors, for_font=True)
        if fc:
            sd["font_color"] = fc

    align = cell.alignment
    if align:
        if align.horizontal in ("center", "right", "left"):
            sd["h_align"] = align.horizontal
        sd["wrap_text"] = bool(align.wrap_text)

    b = cell.border
    if b:
        sd["border_top"]    = _border_style_name(b.top)
        sd["border_bottom"] = _border_style_name(b.bottom)
        sd["border_left"]   = _border_style_name(b.left)
        sd["border_right"]  = _border_style_name(b.right)

    return sd


def _style_fingerprint(sd: dict) -> tuple:
    return tuple((k, sd[k]) for k in sorted(sd))


def _strip_nulls(sd: dict) -> dict:
    return {k: v for k, v in sd.items() if v is not None and v is not False}


# ── Column width helper ───────────────────────────────────────────────────────

def _col_width_px(ws, col_idx: int) -> int | None:
    from openpyxl.utils import get_column_letter
    letter = get_column_letter(col_idx)
    dim = ws.column_dimensions.get(letter)
    if dim and dim.width:
        return max(30, int(dim.width * 7))
    return None


# ── Full mode: sheet → HTML table ────────────────────────────────────────────

def sheet_to_html(ws, no_borders: bool = False, theme_colors=None) -> str:
    from openpyxl.utils import get_column_letter

    merge_spans, merged_skip = _build_merge_map(ws)

    if ws.max_row is None or ws.max_column is None:
        return "_（empty sheet）_"
    max_row, max_col = ws.max_row, ws.max_column
    if max_row == 0 or max_col == 0:
        return "_（empty sheet）_"

    col_widths = []
    for c in range(1, max_col + 1):
        px = _col_width_px(ws, c)
        col_widths.append(f'<col style="width:{px}px">' if px else "<col>")

    lines = ["<table>", "<colgroup>"] + col_widths + ["</colgroup>", "<tbody>"]

    for r in range(1, max_row + 1):
        row_dim = ws.row_dimensions.get(r)
        row_style = ""
        if row_dim and row_dim.height:
            px = int(row_dim.height * 1.333)
            row_style = f' style="height:{px}px"'
        row_parts = [f"  <tr{row_style}>"]
        has_content = False

        for c in range(1, max_col + 1):
            if (r, c) in merged_skip:
                continue
            cell = ws.cell(row=r, column=c)
            value = cell.value
            display = "" if value is None else str(value)
            if display.strip():
                has_content = True

            css = _cell_css(cell, no_borders, theme_colors)
            attrs = f' style="{css}"' if css else ""

            span_r, span_c = merge_spans.get((r, c), (1, 1))
            if span_r > 1:
                attrs += f' rowspan="{span_r}"'
            if span_c > 1:
                attrs += f' colspan="{span_c}"'

            safe = (display
                    .replace("&", "&amp;")
                    .replace("<", "&lt;")
                    .replace(">", "&gt;"))
            row_parts.append(f"    <td{attrs}>{safe}</td>")

        row_parts.append("  </tr>")

        if not has_content:
            has_span = any((r, c) in merged_skip for c in range(1, max_col + 1))
            if not has_span:
                continue

        lines.extend(row_parts)

    lines.extend(["</tbody>", "</table>"])
    return "\n".join(lines)


# ── Data mode: sheet → plain MD table ────────────────────────────────────────

def sheet_to_md_plain(ws) -> str:
    _, merged_skip = _build_merge_map(ws)

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return "_（empty sheet）_"

    rows: list[list[str]] = []
    for r in range(1, max_row + 1):
        row = []
        for c in range(1, max_col + 1):
            if (r, c) in merged_skip:
                row.append("")
                continue
            cell = ws.cell(row=r, column=c)
            val = "" if cell.value is None else str(cell.value)
            row.append(val.replace("|", "\\|").replace("\n", "<br>"))
        rows.append(row)

    if not rows:
        return "_（empty sheet）_"

    header = rows[0]
    n = len(header)
    lines = [
        "| " + " | ".join(header) + " |",
        "|" + "|".join("---" for _ in range(n)) + "|",
    ]
    for row in rows[1:]:
        while len(row) < n:
            row.append("")
        lines.append("| " + " | ".join(row[:n]) + " |")

    return "\n".join(lines)


# ── Guide mode: extract style guide ──────────────────────────────────────────

def extract_style_guide(ws, theme_colors=None) -> dict:
    """Analyze worksheet → reusable style guide dict."""
    from openpyxl.utils import get_column_letter

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return {}

    _, merged_skip = _build_merge_map(ws)

    # Per-cell style extraction
    cell_styles: dict[tuple, dict] = {}
    cell_fps: dict[tuple, tuple] = {}
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if (r, c) in merged_skip:
                continue
            cell = ws.cell(row=r, column=c)
            sd = _extract_style_dict(cell, theme_colors)
            cell_styles[(r, c)] = sd
            cell_fps[(r, c)] = _style_fingerprint(sd)

    fp_counts = Counter(cell_fps.values())

    # Dominant style per row
    row_dom: dict[int, tuple] = {}
    for r in range(1, max_row + 1):
        row_fp_list = [cell_fps[(r, c)] for c in range(1, max_col + 1) if (r, c) in cell_fps]
        if row_fp_list:
            row_dom[r] = Counter(row_fp_list).most_common(1)[0][0]

    # Cluster consecutive rows with same dominant fp
    clusters: list[tuple[int, int, tuple]] = []  # (start_r, end_r, fp)
    prev_fp: tuple | None = None
    start_r = 1
    for r in range(1, max_row + 2):
        cur_fp = row_dom.get(r)
        if cur_fp != prev_fp:
            if prev_fp is not None:
                clusters.append((start_r, r - 1, prev_fp))
            prev_fp = cur_fp
            start_r = r

    # Auto-name fingerprints used in clusters
    cluster_fps = {c[2] for c in clusters}
    most_common_fp = fp_counts.most_common(1)[0][0] if fp_counts else None

    fp_names: dict[tuple, str] = {}
    header_n = label_n = other_n = 0
    for fp_val in cluster_fps:
        sd = dict(fp_val)
        if fp_val == most_common_fp:
            fp_names[fp_val] = "data"
        elif sd.get("bold") and sd.get("bg_color"):
            name = "header" if header_n == 0 else f"header_{header_n + 1}"
            fp_names[fp_val] = name
            header_n += 1
        elif sd.get("bold"):
            name = "label" if label_n == 0 else f"label_{label_n + 1}"
            fp_names[fp_val] = name
            label_n += 1
        else:
            other_n += 1
            fp_names[fp_val] = f"style_{other_n}"

    # Build named_styles (strip null/False values for readability)
    named_styles: dict[str, dict] = {}
    for fp_val, name in fp_names.items():
        named_styles[name] = _strip_nulls(dict(fp_val))

    # Build range_styles from clusters
    range_styles: list[dict] = []
    for start_r, end_r, fp_val in clusters:
        name = fp_names.get(fp_val)
        if name:
            range_styles.append({
                "range": f"A{start_r}:{get_column_letter(max_col)}{end_r}",
                "style": name,
            })

    # Cell overrides: cells that differ from their row's dominant
    cell_overrides: dict[str, dict] = {}
    for (r, c), fp_val in cell_fps.items():
        dom = row_dom.get(r)
        if dom and fp_val != dom:
            sd = _strip_nulls(cell_styles[(r, c)])
            dom_dict = dict(dom)
            # Explicitly mark bg_color/font_color as null when dominant has them but cell doesn't.
            # This tells apply_style_guide to clear those properties instead of inheriting the range style.
            for key in ("bg_color", "font_color"):
                if cell_styles[(r, c)].get(key) is None and dom_dict.get(key) is not None:
                    sd[key] = None
            if sd:
                cell_overrides[f"{get_column_letter(c)}{r}"] = sd

    # Column widths (Excel char units)
    column_widths: dict[str, float] = {}
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        if dim and dim.width:
            column_widths[letter] = round(dim.width, 2)

    # Row heights (pt)
    row_heights: dict[str, float] = {}
    for r in range(1, max_row + 1):
        dim = ws.row_dimensions.get(r)
        if dim and dim.height:
            row_heights[str(r)] = round(dim.height, 2)

    # Merge ranges
    merges = [str(rng) for rng in ws.merged_cells.ranges]

    return {
        "column_widths": column_widths,
        "row_heights": row_heights,
        "named_styles": named_styles,
        "range_styles": range_styles,
        "cell_overrides": cell_overrides,
        "merges": merges,
    }


# ── Embedded image helpers ───────────────────────────────────────────────────

def _detect_images(xlsx_path: Path) -> list[str]:
    """Return list of media filenames inside xl/media/ (images embedded in the workbook)."""
    import zipfile
    with zipfile.ZipFile(str(xlsx_path)) as z:
        return [Path(n).name for n in z.namelist() if n.startswith("xl/media/")]


def extract_images(xlsx_path: Path, out_dir: Path, save_manifest: bool = False) -> Path:
    """
    Extract embedded images from xlsx to out_dir.
    If save_manifest=True, saves the full drawing infrastructure to manifest.json:
      - drawing XMLs  (xl/drawings/*.xml)
      - drawing rels  (xl/drawings/_rels/*.rels)
      - sheet rels    (xl/worksheets/_rels/sheet*.rels that reference drawings)
      - content-type entries for drawings and media
    This makes the manifest self-sufficient — original file not needed for reconstruction.
    Returns out_dir.
    """
    import zipfile

    out_dir.mkdir(parents=True, exist_ok=True)
    manifest: dict = {
        "source": xlsx_path.name,
        "images": [],
        "drawings": {},
        "drawing_rels": {},
        "sheet_rels": {},
        "content_type_overrides": [],
    }

    with zipfile.ZipFile(str(xlsx_path)) as z:
        names = set(z.namelist())

        # Collect sheet rels that reference drawings (need to scan before iterating)
        drawing_sheet_rels: set[str] = set()
        if save_manifest:
            for n in names:
                if re.match(r"xl/worksheets/_rels/sheet\d+\.xml\.rels$", n):
                    if b"/drawing" in z.read(n):
                        drawing_sheet_rels.add(n)

        for name in sorted(names):
            stem = Path(name).name

            if name.startswith("xl/media/"):
                (out_dir / stem).write_bytes(z.read(name))
                manifest["images"].append(stem)
                print(f"  Extracted: {stem}")

            elif save_manifest:
                if "xl/drawings/" in name and name.endswith(".xml") and "_rels" not in name:
                    manifest["drawings"][stem] = z.read(name).decode("utf-8")
                elif "xl/drawings/_rels/" in name and name.endswith(".rels"):
                    manifest["drawing_rels"][stem] = z.read(name).decode("utf-8")
                elif name in drawing_sheet_rels:
                    manifest["sheet_rels"][stem] = z.read(name).decode("utf-8")

        # Save drawing + media content-type entries from [Content_Types].xml
        if save_manifest and "[Content_Types].xml" in names:
            ct = z.read("[Content_Types].xml").decode("utf-8")
            entries = re.findall(r'<Override[^>]*/xl/drawings[^>]*/>', ct)
            entries += re.findall(
                r'<Default[^>]+(?:png|jpeg|jpg|gif|bmp|emf|wmf)[^>]*/>', ct, re.IGNORECASE
            )
            manifest["content_type_overrides"] = entries

    if save_manifest:
        (out_dir / "manifest.json").write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"  Saved: manifest.json  "
              f"({len(manifest['drawings'])} drawings, "
              f"{len(manifest['sheet_rels'])} sheet rels, "
              f"{len(manifest['drawing_rels'])} drawing rels)")

    print(f"Done -> {out_dir} ({len(manifest['images'])} image(s))")
    return out_dir


# ── Auto-screenshot helper ────────────────────────────────────────────────────

def _auto_screenshot(xlsx_path: Path, out_dir: Path) -> list[Path]:
    """Call screenshot_sheets.py logic inline. Returns [] if dependencies missing."""
    try:
        script_dir = Path(__file__).parent
        sys.path.insert(0, str(script_dir))
        from screenshot_sheets import screenshot_xlsx
        screenshots_dir = out_dir / (xlsx_path.stem + "_screenshots")
        return screenshot_xlsx(xlsx_path, screenshots_dir, dpi=150)
    except ImportError:
        return []
    except Exception as e:
        print(f"  Screenshot failed: {e}")
        return []


# ── Config management ─────────────────────────────────────────────────────────

_CONFIG_PATH = Path.home() / ".docling_convert_config.json"


def _load_config() -> dict:
    if _CONFIG_PATH.exists():
        try:
            return json.loads(_CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _save_config(cfg: dict):
    _CONFIG_PATH.write_text(json.dumps(cfg, indent=2), encoding="utf-8")


def _prompt_mode() -> str:
    print()
    print("Choose conversion mode (saved for future runs - pass --reset to change):")
    print("  [1] Extract    - MD tables only. Fast, LLM-readable, no styling.")
    print("  [2] Round-trip - MD + style guide. For reconstructing or restyling Excel.")
    while True:
        try:
            raw = input("  Choice [1/2]: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nCancelled.")
            sys.exit(0)
        if raw == "1":
            return "extract"
        if raw == "2":
            return "roundtrip"
        print("  Enter 1 or 2.")


def _get_mode(reset: bool) -> str:
    cfg = _load_config()
    if not reset and "mode" in cfg:
        print(f"[Mode: {cfg['mode']}]  (pass --reset to change)")
        return cfg["mode"]
    mode = _prompt_mode()
    cfg["mode"] = mode
    _save_config(cfg)
    print(f"[Mode saved: {mode}]")
    return mode


# ── Screenshot request generator ──────────────────────────────────────────────

def _screenshot_requests(sheet_names: list[str], file_name: str) -> str:
    sep = "=" * 60
    lines = [
        sep,
        "STYLE VERIFICATION - SCREENSHOT REQUEST",
        sep,
        "",
        f"Style guide extracted from: {file_name}",
        "Some visual properties (fills, conditional formatting, subtle",
        "colors) cannot be reliably read from file metadata alone.",
        "",
        "To correct the style guide, please screenshot the following",
        "and share here:",
        "",
    ]
    for name in sheet_names:
        lines.append(f"  - Sheet '{name}'  (full view, all rows visible)")
    lines.extend([
        "",
        "Once screenshots are shared, the style guide will be corrected",
        "based on what is visible in the screenshots.",
        sep,
    ])
    return "\n".join(lines)


# ── Main conversion ───────────────────────────────────────────────────────────

def convert(input_path: Path, output_path: Path, mode: str):
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    print(f"Loading: {input_path.name}")
    wb = openpyxl.load_workbook(str(input_path), data_only=True)
    sheets = wb.sheetnames
    print(f"Sheets: {len(sheets)}")

    theme_colors = _resolve_theme_colors(wb)
    if theme_colors:
        print(f"Theme colors: {len(theme_colors)} slots resolved")

    md_parts = []
    for name in sheets:
        ws = wb[name]
        print(f"  {name} ...")
        md = sheet_to_md_plain(ws)
        md_parts.append(f"# Sheet: {name}\n\n{md}\n")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("".join(md_parts), encoding="utf-8")
    print(f"Saved: {output_path} ({output_path.stat().st_size // 1024} KB)")

    if mode == "roundtrip":
        guide: dict = {"version": 1, "sheets": {}}
        for name in sheets:
            ws = wb[name]
            guide["sheets"][name] = extract_style_guide(ws, theme_colors)

        stem = re.sub(r"_data$", "", output_path.stem)
        guide_path = output_path.with_name(stem + "_styles.json")
        guide_path.write_text(
            json.dumps(guide, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"Saved: {guide_path}")

        # Auto-screenshot each sheet for style verification
        print("\nTaking screenshots for style verification ...")
        screenshots = _auto_screenshot(input_path, output_path.parent)
        if screenshots:
            print(f"Screenshots saved: {len(screenshots)} file(s)")
            print()
            print("AI REVIEW STEP")
            print("-" * 40)
            for p in screenshots:
                print(f"  SCREENSHOT: {p}")
            print(f"  STYLES_JSON: {guide_path}")
            print("-" * 40)
            print("AI should now read screenshots and correct styles.json.")
        else:
            print("Screenshots unavailable (install pywin32 + pymupdf to enable).")
            print()
            print(_screenshot_requests(sheets, input_path.name))


def main():
    parser = argparse.ArgumentParser(
        description="Excel → Markdown. Mode chosen once and saved."
    )
    parser.add_argument("input", help="Input .xlsx or .xlsm file")
    parser.add_argument("--out", help="Output .md path (default: <stem>_data.md)")
    parser.add_argument("--reset", action="store_true", help="Forget saved mode and choose again")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    mode = _get_mode(args.reset)

    out = Path(args.out) if args.out else input_path.with_name(input_path.stem + "_data.md")

    # ── Image handling ────────────────────────────────────────────────────────
    images = _detect_images(input_path)
    if images:
        print(f"\nEmbedded images found: {len(images)} file(s)")
        for img in images:
            print(f"  {img}")
        if mode == "roundtrip":
            img_dir = out.parent / (input_path.stem + "_images")
            extract_images(input_path, img_dir, save_manifest=True)
            print(f"  Images + manifest saved to: {img_dir}/")
            print(f"  (Needed later for reconstruct_xlsx_images.py)")
        else:
            print("  Extract mode - images ignored.")
        print()

    # ── Convert ───────────────────────────────────────────────────────────────
    print(f"\nConverting {input_path.name} ...")
    convert(input_path, out, mode)


if __name__ == "__main__":
    main()
