"""
convert_styled_md_to_xlsx.py
Reconstruct Excel (.xlsx) from a styled Markdown file produced by convert_xlsx_styled.py.

Reads # Sheet: Name sections → HTML tables → applies full openpyxl styling:
  - Cell values
  - Background fill colors
  - Font: bold, italic, size, color
  - Merged cells (colspan / rowspan)
  - Column widths
  - Text alignment
  - Borders

Requires:
    pip install openpyxl  (already installed with docling)

Usage:
    python convert_styled_md_to_xlsx.py input_styled.md [--out output.xlsx]
"""
import argparse
import re
import sys
from html.parser import HTMLParser
from pathlib import Path


# ── CSS parser ────────────────────────────────────────────────────────────────

def parse_css(style_str: str) -> dict[str, str]:
    result = {}
    for part in style_str.split(";"):
        part = part.strip()
        if ":" not in part:
            continue
        k, _, v = part.partition(":")
        result[k.strip().lower()] = v.strip()
    return result


def css_color(val: str) -> str | None:
    """Extract 6-char hex from #RGB or #RRGGBB."""
    val = val.strip()
    m = re.match(r"#([0-9a-fA-F]{6})", val)
    if m:
        return m.group(1).upper()
    m = re.match(r"#([0-9a-fA-F]{3})$", val)
    if m:
        c = m.group(1)
        return (c[0]*2 + c[1]*2 + c[2]*2).upper()
    return None


def css_border_style(val: str) -> str | None:
    val = val.lower()
    if "2px" in val or "medium" in val:
        return "medium"
    if "3px" in val or "thick" in val:
        return "thick"
    if "double" in val:
        return "double"
    if "dashed" in val:
        return "dashed"
    if "dotted" in val:
        return "dotted"
    if "1px" in val or "solid" in val or "thin" in val or "hair" in val:
        return "thin"
    return None


def css_border_color(val: str) -> str | None:
    m = re.search(r"#([0-9a-fA-F]{6})", val)
    return m.group(1).upper() if m else None


# ── HTML table parser ─────────────────────────────────────────────────────────

class TableParser(HTMLParser):
    """Parse a single <table> block into a list of rows with cell metadata."""

    def __init__(self):
        super().__init__()
        self.col_widths: list[int | None] = []
        self.rows: list[list[dict]] = []
        self.row_heights: list[int | None] = []  # px per row
        self._cur_row: list[dict] | None = None
        self._cur_cell: dict | None = None
        self._in_cell = False

    def handle_starttag(self, tag, attrs):
        attrs = dict(attrs)
        if tag == "col":
            style = parse_css(attrs.get("style", ""))
            w = style.get("width", "")
            m = re.match(r"(\d+)px", w)
            self.col_widths.append(int(m.group(1)) if m else None)
        elif tag == "tr":
            self._cur_row = []
            style = parse_css(attrs.get("style", ""))
            h_str = style.get("height", "")
            m = re.match(r"(\d+)px", h_str)
            self.row_heights.append(int(m.group(1)) if m else None)
        elif tag == "td" or tag == "th":
            style = parse_css(attrs.get("style", ""))
            self._cur_cell = {
                "style": style,
                "rowspan": int(attrs.get("rowspan", 1)),
                "colspan": int(attrs.get("colspan", 1)),
                "text": "",
            }
            self._in_cell = True

    def handle_endtag(self, tag):
        if tag == "tr":
            if self._cur_row is not None:
                self.rows.append(self._cur_row)
            self._cur_row = None
        elif tag in ("td", "th"):
            if self._cur_row is not None and self._cur_cell is not None:
                self._cur_row.append(self._cur_cell)
            self._cur_cell = None
            self._in_cell = False

    def handle_data(self, data):
        if self._in_cell and self._cur_cell is not None:
            self._cur_cell["text"] += data

    def handle_entityref(self, name):
        mapping = {"amp": "&", "lt": "<", "gt": ">", "nbsp": " "}
        if self._in_cell and self._cur_cell is not None:
            self._cur_cell["text"] += mapping.get(name, "")

    def handle_charref(self, name):
        try:
            ch = chr(int(name[1:], 16) if name.startswith("x") else int(name))
        except Exception:
            ch = ""
        if self._in_cell and self._cur_cell is not None:
            self._cur_cell["text"] += ch


# ── Apply style to openpyxl cell ─────────────────────────────────────────────

def apply_style(cell, style: dict, no_fill_colors=False):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    def _argb(hex6: str) -> str:
        return hex6 if len(hex6) == 8 else "FF" + hex6

    # Fill
    bg = css_color(style.get("background-color", ""))
    if bg and not no_fill_colors:
        cell.fill = PatternFill(fill_type="solid", fgColor=_argb(bg))

    # Font
    bold   = style.get("font-weight") == "bold"
    italic = style.get("font-style") == "italic"
    fc     = css_color(style.get("color", ""))
    fsize  = None
    m = re.match(r"(\d+(?:\.\d+)?)", style.get("font-size", ""))
    if m:
        fsize = float(m.group(1))
    cell.font = Font(
        bold=bold,
        italic=italic,
        size=fsize,
        color=_argb(fc) if fc else "FF000000",
    )

    # Alignment
    ha = style.get("text-align")
    wrap = style.get("white-space") in ("pre-wrap", "pre", "wrap")
    if ha in ("center", "right", "left"):
        cell.alignment = Alignment(horizontal=ha, wrap_text=wrap)
    elif wrap:
        cell.alignment = Alignment(wrap_text=True)

    # Borders
    def side(key):
        val = style.get(key) or style.get("border")
        if not val:
            return Side()
        bs = css_border_style(val)
        bc = css_border_color(val)
        if bs:
            return Side(border_style=bs, color=bc or "000000")
        return Side()

    cell.border = Border(
        top=side("border-top"),
        bottom=side("border-bottom"),
        left=side("border-left"),
        right=side("border-right"),
    )


# ── Build worksheet from parsed table ─────────────────────────────────────────

def table_to_sheet(ws, parser: TableParser):
    from openpyxl.utils import get_column_letter

    # Apply column widths
    for idx, px in enumerate(parser.col_widths, 1):
        if px:
            ws.column_dimensions[get_column_letter(idx)].width = px / 7.0

    # Apply row heights (pt = px / 1.333)
    for r_idx, px in enumerate(parser.row_heights, 1):
        if px:
            ws.row_dimensions[r_idx].height = px / 1.333

    # Track which (row, col) positions are occupied by rowspan/colspan
    occupied: dict[tuple[int, int], bool] = {}
    merges: list[tuple[int, int, int, int]] = []  # (r1, c1, r2, c2)

    for r_idx, row_cells in enumerate(parser.rows, 1):
        c_idx = 1
        for cell_data in row_cells:
            # Advance past occupied cells
            while occupied.get((r_idx, c_idx)):
                c_idx += 1

            rs = cell_data["rowspan"]
            cs = cell_data["colspan"]
            text = cell_data["text"].strip()
            style = cell_data["style"]

            # Write value — try numeric conversion
            value: object = text
            if text:
                try:
                    value = int(text)
                except ValueError:
                    try:
                        value = float(text)
                    except ValueError:
                        value = text

            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            apply_style(cell, style)

            # Mark occupied
            for dr in range(rs):
                for dc in range(cs):
                    if dr > 0 or dc > 0:
                        occupied[(r_idx + dr, c_idx + dc)] = True

            if rs > 1 or cs > 1:
                merges.append((r_idx, c_idx, r_idx + rs - 1, c_idx + cs - 1))

            c_idx += cs

    # Apply merges after all cells written
    for r1, c1, r2, c2 in merges:
        ws.merge_cells(
            start_row=r1, start_column=c1,
            end_row=r2, end_column=c2,
        )


# ── MD parser ────────────────────────────────────────────────────────────────

def parse_styled_md(text: str) -> list[tuple[str, str]]:
    """Split by '# Sheet: Name' → list of (sheet_name, html_content)."""
    parts = re.split(r"^# Sheet: (.+)$", text, flags=re.MULTILINE)
    pairs = []
    for name, content in zip(parts[1::2], parts[2::2]):
        pairs.append((name.strip(), content.strip()))
    return pairs


# ── Main ──────────────────────────────────────────────────────────────────────

def convert(input_path: Path, output_path: Path):
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    text = input_path.read_text(encoding="utf-8")
    sheets = parse_styled_md(text)
    print(f"Sheets found: {len(sheets)}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for name, html in sheets:
        parser = TableParser()
        parser.feed(html)

        safe_name = name[:31]
        ws = wb.create_sheet(title=safe_name)
        table_to_sheet(ws, parser)

        n_rows = len(parser.rows)
        print(f"  {name}: {n_rows} rows, {len(parser.col_widths)} cols")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"\nSaved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Styled Markdown → Excel with full styling")
    parser.add_argument("input", help="Input _styled.md file")
    parser.add_argument("--out", help="Output .xlsx path (default: <stem>.xlsx)")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    stem = re.sub(r"_styled$", "", input_path.stem)
    out = Path(args.out) if args.out else input_path.with_name(stem + "_reconstructed.xlsx")
    print(f"Converting {input_path.name} → {out.name}...")
    convert(input_path, out)


if __name__ == "__main__":
    main()
