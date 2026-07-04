"""Markdown-to-Excel Converter — Local CLI Tool.

Converts Markdown files to Excel (.xlsx) without sending content to any AI server.
All processing is done entirely locally.

Features:
  - Auto-detect all markdown tables in a file -> one sheet per table
  - Export by section heading: --section "## 6. Section Name"
  - Export by row range within a table: --from-row N --to-row M
  - Support multiple tables in the same section

Usage:
    python convert_md_to_excel.py --source <file.md>
    python convert_md_to_excel.py --source <file.md> --output <out.xlsx>
    python convert_md_to_excel.py --source <file.md> --section "## 6. Risk Log"
    python convert_md_to_excel.py --source <file.md> --section "## 6. Risk Log" --from-row 5 --to-row 20
    python convert_md_to_excel.py --source <file.md> --from-row 5 --to-row 20
    python convert_md_to_excel.py --source <file.md> --list-sections

Requires: openpyxl  (pip install openpyxl)
"""
from __future__ import annotations

import argparse
import logging
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Set, Tuple

# Force UTF-8 stdout/stderr so non-ASCII (— etc.) works on Windows cp932.
for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
    except Exception:
        pass

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write(
        "ERROR: openpyxl is required. Install it with:\n"
        "    pip install openpyxl\n"
    )
    sys.exit(2)


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class MarkdownTable:
    """A single markdown table extracted from a document."""

    header: List[str]                # Column headers
    rows: List[List[str]]            # Data rows (excluding header)
    section: str = ""                # Section heading this table belongs to
    table_index: int = 0             # 0-based index among tables in the same section
    start_line: int = 0              # Line number in source file (1-based)


@dataclass
class Section:
    """A section (heading + content) in the markdown document."""

    heading: str                     # e.g. "## 6. Risk Log"
    level: int                       # 1=H1, 2=H2, etc.
    start_line: int                  # Line number where heading appears (1-based)
    end_line: int = -1               # Line number where section ends (exclusive)
    tables: List[MarkdownTable] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Markdown parsing
# ---------------------------------------------------------------------------


def _heading_level(line: str) -> int:
    """Return heading level (1-6) or 0 if not a heading."""
    m = re.match(r"^(#{1,6})\s", line)
    return len(m.group(1)) if m else 0


def _is_table_separator(line: str) -> bool:
    """Return True if line is a markdown table separator (e.g. |---|---|)."""
    stripped = line.strip()
    if not stripped.startswith("|") and "|" not in stripped:
        return False
    return bool(re.match(r"^[\|\s\-:]+$", stripped))


def _parse_table_row(line: str) -> List[str]:
    """Parse a markdown table row into cells, stripping whitespace."""
    cells = line.strip().split("|")
    if cells and cells[0].strip() == "":
        cells = cells[1:]
    if cells and cells[-1].strip() == "":
        cells = cells[:-1]
    return [c.strip() for c in cells]


def _normalize_heading(heading: str) -> str:
    """Normalize heading for comparison: lowercase, collapse whitespace."""
    return re.sub(r"\s+", " ", heading.strip()).lower()


def parse_markdown(source_path: Path) -> Tuple[List[Section], List[MarkdownTable]]:
    """Parse a markdown file into sections and tables.

    Returns (sections, orphan_tables) where orphan_tables are those that do
    not belong to any section.
    """
    lines = source_path.read_text(encoding="utf-8").splitlines()
    sections: List[Section] = []
    current_section: Optional[Section] = None
    orphan_tables: List[MarkdownTable] = []

    i = 0
    while i < len(lines):
        line = lines[i]
        level = _heading_level(line)

        if level > 0:
            if current_section is not None:
                current_section.end_line = i + 1  # exclusive, 1-based
            current_section = Section(
                heading=line.strip(),
                level=level,
                start_line=i + 1,
            )
            sections.append(current_section)
            i += 1
            continue

        # Detect start of a table: pipe row followed by separator row.
        if "|" in line and i + 1 < len(lines) and _is_table_separator(lines[i + 1]):
            header_cells = _parse_table_row(line)
            table_rows: List[List[str]] = []
            i += 2  # skip separator line

            while i < len(lines):
                tline = lines[i]
                if "|" not in tline:
                    break
                table_rows.append(_parse_table_row(tline))
                i += 1

            section_name = current_section.heading if current_section else ""
            table_idx = len(current_section.tables) if current_section else len(orphan_tables)

            table = MarkdownTable(
                header=header_cells,
                rows=table_rows,
                section=section_name,
                table_index=table_idx,
                start_line=i - len(table_rows) - 2 + 1,
            )

            if current_section is not None:
                current_section.tables.append(table)
            else:
                orphan_tables.append(table)
            continue

        i += 1

    if current_section is not None:
        current_section.end_line = len(lines) + 1

    return sections, orphan_tables


# ---------------------------------------------------------------------------
# Filtering helpers
# ---------------------------------------------------------------------------


def find_section(sections: List[Section], section_query: str) -> Optional[Section]:
    """Find a section by heading. Exact match first, then partial (case-insensitive)."""
    query_norm = _normalize_heading(section_query)
    for sec in sections:
        if _normalize_heading(sec.heading) == query_norm:
            return sec
    for sec in sections:
        if query_norm in _normalize_heading(sec.heading):
            return sec
    return None


def filter_rows_by_range(
    tables: List[MarkdownTable],
    from_row: Optional[int],
    to_row: Optional[int],
) -> List[MarkdownTable]:
    """Filter table rows by 1-based row numbers (data rows only, not header).

    from_row=1 means first data row. Inclusive on both ends.
    """
    if from_row is None and to_row is None:
        return tables

    result = []
    for table in tables:
        total = len(table.rows)
        fr = (from_row - 1) if from_row else 0          # convert to 0-based
        tr = to_row if to_row else total                  # exclusive end

        sliced = table.rows[fr:tr]
        result.append(MarkdownTable(
            header=table.header,
            rows=sliced,
            section=table.section,
            table_index=table.table_index,
            start_line=table.start_line,
        ))

    return result


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------


def _make_sheet_name(table: MarkdownTable, used_names: Set[str]) -> str:
    """Generate a unique, valid Excel sheet name (max 31 chars)."""
    base = re.sub(r"^#+\s*", "", table.section).strip() if table.section else "Table"
    if table.table_index > 0:
        base = f"{base} ({table.table_index + 1})"
    base = re.sub(r"[\\/*?\[\]:]", "_", base)[:31]
    name = base
    counter = 2
    while name in used_names:
        suffix = f" ({counter})"
        name = base[: 31 - len(suffix)] + suffix
        counter += 1
    used_names.add(name)
    return name


def _style_header_row(ws, row_idx: int, ncols: int) -> None:
    """Apply bold + light blue fill to the header row."""
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _style_data_row(ws, row_idx: int, ncols: int, shade: bool) -> None:
    """Apply alternating row shading."""
    fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") if shade else None
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _auto_col_width(ws, max_width: int = 60) -> None:
    """Auto-size columns based on content, capped at max_width."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                for part in str(cell.value).splitlines():
                    max_len = max(max_len, len(part))
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


def _add_thin_border(ws, nrows: int, ncols: int) -> None:
    """Add a thin border around all cells in the data range."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, nrows + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = border


def write_excel(
    tables: List[MarkdownTable],
    output_path: Path,
    source_name: str = "",
) -> int:
    """Write tables to an Excel workbook (one table per sheet).

    Returns number of sheets written.
    """
    if not tables:
        logger.warning("No tables to write.")
        return 0

    wb = openpyxl.Workbook()
    if wb.active:
        wb.remove(wb.active)

    used_names: Set[str] = set()

    for table in tables:
        if not table.header:
            logger.warning("Skipping table with no header (section=%s)", table.section)
            continue

        sheet_name = _make_sheet_name(table, used_names)
        ws = wb.create_sheet(title=sheet_name)
        ws.freeze_panes = "A2"

        ncols = len(table.header)

        for col_idx, cell_val in enumerate(table.header, start=1):
            ws.cell(row=1, column=col_idx, value=cell_val)
        _style_header_row(ws, 1, ncols)

        for row_idx, row in enumerate(table.rows, start=2):
            for col_idx in range(ncols):
                val = row[col_idx] if col_idx < len(row) else ""
                ws.cell(row=row_idx, column=col_idx + 1, value=val)
            _style_data_row(ws, row_idx, ncols, shade=(row_idx % 2 == 0))

        total_rows = 1 + len(table.rows)
        _add_thin_border(ws, total_rows, ncols)
        _auto_col_width(ws)

        logger.info("Sheet '%s': %d columns, %d data rows", sheet_name, ncols, len(table.rows))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Saved: %s", output_path)

    return len(wb.sheetnames)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def resolve_output_path(source_path: Path, output_arg: Optional[str]) -> Path:
    """Resolve output .xlsx path from --output arg or default (next to source)."""
    if output_arg:
        p = Path(output_arg)
        if not p.suffix:
            p = p.with_suffix(".xlsx")
        return p
    return source_path.with_suffix(".xlsx")


def list_sections(source_path: Path) -> None:
    """Print all sections and table counts to stdout."""
    sections, orphan_tables = parse_markdown(source_path)
    print(f"\nSections in: {source_path.name}")
    print("-" * 60)
    for sec in sections:
        print(f"  Line {sec.start_line:4d}: {sec.heading}  ({len(sec.tables)} table(s))")
    if orphan_tables:
        print(f"\n  Orphan tables (no section): {len(orphan_tables)}")
    print()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Convert Markdown tables to Excel — runs fully locally.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--source", required=True, help="Path to source .md file")
    parser.add_argument("--output", default=None,
                        help="Output .xlsx path (default: same dir as source)")
    parser.add_argument("--section", default=None,
                        help='Export only tables under this heading (e.g. "## 6. Risk Log")')
    parser.add_argument("--from-row", type=int, default=None,
                        help="Start data row number (1-based, inclusive). Counts only data rows, not header.")
    parser.add_argument("--to-row", type=int, default=None,
                        help="End data row number (1-based, inclusive).")
    parser.add_argument("--list-sections", action="store_true",
                        help="List all sections and table counts, then exit.")
    parser.add_argument("-v", "--verbose", action="store_true", help="Enable DEBUG logging")

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    source_path = Path(args.source)
    if not source_path.exists():
        logger.error("Source file not found: %s", source_path)
        return 1
    if source_path.suffix.lower() not in (".md", ".markdown"):
        logger.error("Source must be a Markdown file (.md, .markdown): %s", source_path)
        return 1

    logger.info("Parsing: %s", source_path)
    sections, orphan_tables = parse_markdown(source_path)

    if args.list_sections:
        list_sections(source_path)
        return 0

    all_tables: List[MarkdownTable] = []

    if args.section:
        sec = find_section(sections, args.section)
        if sec is None:
            logger.error("Section not found: '%s'", args.section)
            logger.info("Available sections:")
            for s in sections:
                logger.info("  %s", s.heading)
            return 1
        logger.info("Found section: %s (%d table(s))", sec.heading, len(sec.tables))
        all_tables = list(sec.tables)
    else:
        for sec in sections:
            all_tables.extend(sec.tables)
        all_tables.extend(orphan_tables)

    if not all_tables:
        suffix = f" under section '{args.section}'" if args.section else ""
        logger.warning("No tables found in source%s.", suffix)
        return 0

    logger.info("Total tables collected: %d", len(all_tables))

    if args.from_row is not None or args.to_row is not None:
        all_tables = filter_rows_by_range(all_tables, args.from_row, args.to_row)
        logger.info(
            "After row filter [%s:%s]: %d table(s)",
            args.from_row or "start", args.to_row or "end", len(all_tables),
        )

    output_path = resolve_output_path(source_path, args.output)

    n_sheets = write_excel(all_tables, output_path, source_name=source_path.name)

    if n_sheets == 0:
        logger.warning("No sheets were written.")
        return 1

    print(f"\nDone. {n_sheets} sheet(s) written to: {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
