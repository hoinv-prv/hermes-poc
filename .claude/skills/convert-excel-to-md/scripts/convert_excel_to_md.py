"""
convert_excel_to_md.py — Excel → Markdown converter for the AIWS skill `convert-excel-to-md`.

Usage:
    python convert_excel_to_md.py --source <path> --output <dir> [options]

Options:
    --source <path>     Excel file (.xlsx / .xls) or directory containing Excel files.
    --output <dir>      Root output directory. Sheet files are created inside it.
    --keep-all-rows     Preserve all rows, including trailing empty ones (default: skip).
    --clean             Delete output directory before conversion.
    -v, --verbose       Enable DEBUG-level logging to stderr.

Output layout:
    Single file  →  <output>/<safe_sheet_name>.md   (one file per sheet)
    Directory    →  <output>/<file_stem>/<safe_sheet_name>.md

Windows note:
    stdout / stderr are forced to UTF-8 at startup so Japanese filenames and cell
    content are printed correctly even when the console codepage is cp932 (Shift-JIS).

Dependencies:
    openpyxl  (install via:  pip install openpyxl  or  pip install -r requirements.txt)
"""

import sys
import io
import argparse
import logging
import re
import shutil
from pathlib import Path

# Force UTF-8 on stdout/stderr — required on Windows cp932 / Shift-JIS environments.
# Must happen before any print() or logging output.
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    print(
        "Error: openpyxl not installed.\n"
        "Run:  pip install openpyxl\n"
        "  or: pip install -r .claude/skills/convert-excel-to-md/requirements.txt",
        file=sys.stderr,
    )
    sys.exit(1)


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

def setup_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s", stream=sys.stderr)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_INVALID_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|]')
_WHITESPACE_RUN = re.compile(r"\s+")
_MAX_STEM_LEN = 60  # keep paths short enough for Windows MAX_PATH


def sanitize_filename(name: str) -> str:
    """Convert a sheet name to a safe filename stem (no extension)."""
    safe = _INVALID_FILENAME_CHARS.sub("_", name)
    safe = _WHITESPACE_RUN.sub("_", safe.strip())
    safe = safe[:_MAX_STEM_LEN]
    return safe if safe else "sheet"


def cell_text(cell) -> str:
    """Return cell value as a Markdown-safe string."""
    if cell.value is None:
        return ""
    text = str(cell.value)
    # Escape pipe so it doesn't break the table, collapse newlines
    return text.replace("|", "\\|").replace("\r\n", " ").replace("\n", " ").replace("\r", " ")


def is_row_empty(row) -> bool:
    return all(c.value is None for c in row)


def sheet_to_markdown(ws, keep_all_rows: bool) -> str:
    """Render a worksheet as a GFM (pipe-table) Markdown string.

    Returns an empty string if the sheet contains no data rows.
    """
    rows = list(ws.iter_rows())
    if not rows:
        return ""

    if not keep_all_rows:
        # Drop trailing all-empty rows before we touch the header
        while rows and is_row_empty(rows[-1]):
            rows.pop()

    if not rows:
        return ""

    col_count = len(rows[0])
    lines: list[str] = []

    # Header row (first row of the sheet)
    header_cells = [cell_text(c) for c in rows[0]]
    lines.append("| " + " | ".join(header_cells) + " |")
    lines.append("| " + " | ".join(["---"] * col_count) + " |")

    # Data rows
    for row in rows[1:]:
        if not keep_all_rows and is_row_empty(row):
            continue
        cells = [cell_text(c) for c in row]
        # Pad shorter rows so column count stays consistent
        if len(cells) < col_count:
            cells.extend([""] * (col_count - len(cells)))
        lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Core conversion
# ---------------------------------------------------------------------------

def convert_file(xlsx_path: Path, output_dir: Path, keep_all_rows: bool) -> int:
    """Convert all sheets in *xlsx_path* into .md files under *output_dir*.

    Returns the number of non-empty sheets written.
    """
    log = logging.getLogger(__name__)
    log.info("Converting: %s", xlsx_path)

    # Pass str(path) so openpyxl handles non-ASCII paths (Japanese, etc.) correctly.
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as exc:
        log.error("Cannot open %s — %s", xlsx_path, exc)
        raise

    output_dir.mkdir(parents=True, exist_ok=True)
    written = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        md_body = sheet_to_markdown(ws, keep_all_rows)

        if not md_body:
            log.debug("  Skipping empty sheet: %s", sheet_name)
            continue

        safe_stem = sanitize_filename(sheet_name)
        out_file = output_dir / f"{safe_stem}.md"

        # H1 heading = original sheet name (not sanitized) for readability
        content = f"# {sheet_name}\n\n{md_body}\n"
        out_file.write_text(content, encoding="utf-8")
        log.info("  Written: %s", out_file)
        written += 1

    wb.close()
    return written


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert Excel (.xlsx / .xls) files to Markdown tables.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--source", required=True, metavar="PATH",
                        help="Excel file or directory")
    parser.add_argument("--output", required=True, metavar="DIR",
                        help="Root output directory")
    parser.add_argument("--keep-all-rows", action="store_true",
                        help="Preserve trailing empty rows (default: skip)")
    parser.add_argument("--clean", action="store_true",
                        help="Delete output directory before conversion")
    parser.add_argument("-v", "--verbose", action="store_true",
                        help="Enable DEBUG logging")
    args = parser.parse_args()

    setup_logging(args.verbose)
    log = logging.getLogger(__name__)

    source = Path(args.source)
    output_root = Path(args.output)

    if not source.exists():
        log.error("Source not found: %s", source)
        sys.exit(1)

    if args.clean and output_root.exists():
        log.info("Cleaning output directory: %s", output_root)
        shutil.rmtree(str(output_root))

    excel_exts = {".xlsx", ".xls"}
    total_files = 0
    total_sheets = 0

    if source.is_file():
        if source.suffix.lower() not in excel_exts:
            log.error("Source must be an .xlsx/.xls file or a directory: %s", source)
            sys.exit(1)

        # Single-file mode: sheets go directly into output_root
        total_sheets = convert_file(source, output_root, args.keep_all_rows)
        total_files = 1

    elif source.is_dir():
        excel_files = sorted(
            p for p in source.rglob("*")
            if p.is_file() and p.suffix.lower() in excel_exts
        )
        if not excel_files:
            log.warning("No Excel files found in: %s", source)
            sys.exit(0)

        for xlsx_path in excel_files:
            # Mirror the sub-path from source root; each file gets its own subfolder.
            rel = xlsx_path.relative_to(source)
            file_out_dir = output_root / rel.parent / xlsx_path.stem
            count = convert_file(xlsx_path, file_out_dir, args.keep_all_rows)
            total_files += 1
            total_sheets += count

    else:
        log.error("Source must be an Excel file or a directory: %s", source)
        sys.exit(1)

    print(
        f"Done: {total_files} file(s), {total_sheets} sheet(s) → {output_root}",
        flush=True,
    )


if __name__ == "__main__":
    main()
