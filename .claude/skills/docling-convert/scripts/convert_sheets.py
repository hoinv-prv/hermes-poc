"""
convert_sheets.py
Split an Excel workbook (xlsx/xlsm) into per-sheet files, convert each with
docling, and merge into a single Markdown file with # Sheet: Name headers.

Usage:
    python convert_sheets.py <input_xlsx_or_xlsm> [--out <output_md>]

If --out is omitted, writes to <input_stem>_sheets.md alongside the input.

Note: Images embedded in sheets show as <!-- image --> placeholders in the
per-sheet MD because openpyxl does not faithfully round-trip Excel drawings.
To extract images from a specific sheet, use extract_sheet_image.py with
the ORIGINAL file's single-pass docling MD (see docling-convert skill).
"""
import argparse
import shutil
import sys
import tempfile
from copy import deepcopy
from pathlib import Path

import openpyxl
from docling.document_converter import DocumentConverter


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input", help="Path to .xlsx or .xlsm file")
    parser.add_argument("--out", help="Output markdown path (default: <stem>_sheets.md)")
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        print(f"Error: file not found: {src}")
        sys.exit(1)

    out_md = Path(args.out) if args.out else src.with_name(src.stem + "_sheets.md")

    # --- Load workbook ---
    print(f"Loading: {src.name}")
    wb = openpyxl.load_workbook(str(src), keep_vba=False, data_only=True)
    sheet_names = wb.sheetnames
    print(f"Sheets: {len(sheet_names)}")

    # --- Split into per-sheet xlsx files ---
    tmpdir = Path(tempfile.mkdtemp(prefix="docling_sheets_"))
    sheet_files = []

    for sheet_name in sheet_names:
        ws_src = wb[sheet_name]
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        ws_new.title = sheet_name[:31]

        for row in ws_src.iter_rows():
            for cell in row:
                ws_new.cell(row=cell.row, column=cell.column, value=cell.value)

        for col_dim in ws_src.column_dimensions.values():
            ws_new.column_dimensions[col_dim.index].width = col_dim.width

        for img in ws_src._images:
            try:
                ws_new.add_image(deepcopy(img))
            except Exception:
                pass

        safe_name = "".join(c if c.isalnum() or c in "-_." else "_" for c in sheet_name)
        out_path = tmpdir / f"{safe_name}.xlsx"
        wb_new.save(str(out_path))
        sheet_files.append((sheet_name, out_path))

    # --- Convert all sheets (model loads once) ---
    print("Converting sheets with docling...")
    converter = DocumentConverter()
    results = []

    for sheet_name, xlsx_path in sheet_files:
        print(f"  {sheet_name} ...", end=" ", flush=True)
        try:
            result = converter.convert(str(xlsx_path))
            md = result.document.export_to_markdown()
            results.append((sheet_name, md, None))
            print(f"OK ({md.count(chr(10))} lines)")
        except Exception as e:
            results.append((sheet_name, "", str(e)))
            print(f"ERROR: {e}")

    # --- Merge ---
    parts = []
    for sheet_name, md, err in results:
        parts.append(f"# Sheet: {sheet_name}\n")
        if err:
            parts.append(f"_Conversion error: {err}_\n\n")
        elif md.strip():
            parts.append(md.strip() + "\n\n")
        else:
            parts.append("_(empty sheet)_\n\n")

    out_md.parent.mkdir(parents=True, exist_ok=True)
    with open(out_md, "w", encoding="utf-8") as f:
        f.write("".join(parts))

    shutil.rmtree(tmpdir)

    ok = sum(1 for _, _, e in results if e is None)
    print(f"\nDone. {ok}/{len(results)} sheets → {out_md}")


if __name__ == "__main__":
    main()
