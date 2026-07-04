"""
apply_style_guide.py
Apply a _styles.json guide to a _data.md file → styled .xlsx.

Use cases:
  1. Reconstruct original Excel from round-trip pair:
       python apply_style_guide.py original_data.md original_styles.json
  2. Apply a template to new content:
       python apply_style_guide.py new_report_data.md company_styles.json --out new_report.xlsx

Requires:
    pip install openpyxl

Usage:
    python apply_style_guide.py data.md styles.json [--out output.xlsx]
"""
import argparse
import json
import re
import sys
from pathlib import Path


# ── Parse data.md ─────────────────────────────────────────────────────────────

def parse_data_md(text: str) -> list[tuple[str, list[list[str]]]]:
    """Split # Sheet: sections → list of (name, rows) where rows is list of str lists."""
    parts = re.split(r"^# Sheet: (.+)$", text, flags=re.MULTILINE)
    result = []
    for name, content in zip(parts[1::2], parts[2::2]):
        rows: list[list[str]] = []
        for line in content.strip().splitlines():
            line = line.strip()
            if not line or re.match(r"^\|([\s:]*-+[\s:-]*\|)+$", line):
                continue  # skip separator rows (must contain actual dashes)
            if line.startswith("|"):
                cells = [c.strip() for c in line[1:].rstrip("|").split("|")]
                rows.append(cells)
        result.append((name.rstrip("\r"), rows))  # preserve trailing spaces (e.g. "変更履歴 ")
    return result


# ── Apply named style dict → openpyxl cell ────────────────────────────────────

def apply_named_style(cell, sd: dict):
    """Apply a style dict (from styles.json) to an openpyxl cell."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    def _argb(hex6: str) -> str:
        return hex6 if len(hex6) == 8 else "FF" + hex6

    # bg_color: explicit None = clear fill (override cancels range style); missing key = leave as-is
    if "bg_color" in sd:
        if sd["bg_color"] is None:
            cell.fill = PatternFill(fill_type=None)
        else:
            cell.fill = PatternFill(fill_type="solid", fgColor=_argb(sd["bg_color"]))

    font_kw: dict = {}
    if sd.get("bold"):
        font_kw["bold"] = True
    if sd.get("italic"):
        font_kw["italic"] = True
    if sd.get("font_size"):
        font_kw["size"] = sd["font_size"]
    # font_color: explicit None = default/auto color; missing key = leave as-is
    fc = sd.get("font_color") if "font_color" not in sd else sd["font_color"]
    if fc:
        font_kw["color"] = _argb(fc)
    if font_kw:
        cell.font = Font(**font_kw)

    align_kw: dict = {}
    ha = sd.get("h_align")
    if ha:
        align_kw["horizontal"] = ha
    if sd.get("wrap_text"):
        align_kw["wrap_text"] = True
    if align_kw:
        cell.alignment = Alignment(**align_kw)

    def side(name):
        if name:
            return Side(border_style=name, color="000000")
        return Side()

    bt = sd.get("border_top")
    bb = sd.get("border_bottom")
    bl = sd.get("border_left")
    br = sd.get("border_right")
    if any([bt, bb, bl, br]):
        cell.border = Border(
            top=side(bt), bottom=side(bb), left=side(bl), right=side(br)
        )


# ── Coordinate helpers ────────────────────────────────────────────────────────

def _range_coords(range_str: str) -> tuple[int, int, int, int]:
    """'A1:Z50' → (min_row, min_col, max_row, max_col)."""
    from openpyxl.utils import column_index_from_string
    m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", range_str.upper())
    if not m:
        return 1, 1, 1, 1
    return (
        int(m.group(2)), column_index_from_string(m.group(1)),
        int(m.group(4)), column_index_from_string(m.group(3)),
    )


def _cell_coords(addr: str) -> tuple[int, int]:
    """'B3' → (row=3, col=2)."""
    from openpyxl.utils import column_index_from_string
    m = re.match(r"([A-Z]+)(\d+)", addr.upper())
    if not m:
        return 1, 1
    return int(m.group(2)), column_index_from_string(m.group(1))


# ── Build worksheet from data + guide ─────────────────────────────────────────

def apply_guide_to_sheet(ws, rows: list[list[str]], guide: dict):
    from openpyxl.utils import get_column_letter

    named_styles  = guide.get("named_styles", {})
    range_styles  = guide.get("range_styles", [])
    cell_overrides = guide.get("cell_overrides", {})
    merges        = guide.get("merges", [])
    col_widths    = guide.get("column_widths", {})
    row_heights   = guide.get("row_heights", {})

    actual_rows = len(rows)
    actual_cols = max((len(r) for r in rows), default=0)

    # Write data values
    for r_idx, row in enumerate(rows, 1):
        for c_idx, raw in enumerate(row, 1):
            text = raw.strip().replace("\\|", "|").replace("<br>", "\n")
            value: object = text
            if text:
                try:
                    value = int(text)
                except ValueError:
                    try:
                        value = float(text)
                    except ValueError:
                        value = text
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Apply range styles (clamp to actual data extent)
    for rs in range_styles:
        style_name = rs.get("style", "")
        sd = named_styles.get(style_name)
        if not sd:
            continue
        r1, c1, r2, c2 = _range_coords(rs["range"])
        r2 = min(r2, actual_rows)
        c2 = min(c2, actual_cols)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                apply_named_style(ws.cell(row=r, column=c), sd)

    # Apply cell-level overrides
    for addr, sd in cell_overrides.items():
        r, c = _cell_coords(addr)
        if r <= actual_rows and c <= actual_cols:
            apply_named_style(ws.cell(row=r, column=c), sd)

    # Column widths (Excel char units stored in JSON)
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    # Row heights (pt stored in JSON)
    for r_str, height in row_heights.items():
        ws.row_dimensions[int(r_str)].height = height

    # Merge cells (best-effort — skip if range exceeds data)
    for merge_str in merges:
        try:
            ws.merge_cells(merge_str)
        except Exception:
            pass


# ── Main ──────────────────────────────────────────────────────────────────────

def convert(data_path: Path, guide_path: Path, output_path: Path):
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    text  = data_path.read_text(encoding="utf-8")
    guide = json.loads(guide_path.read_text(encoding="utf-8"))

    sheets = parse_data_md(text)
    print(f"Sheets found: {len(sheets)}")

    guide_sheets = guide.get("sheets", {})

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for name, rows in sheets:
        ws = wb.create_sheet(title=name[:31])
        sheet_guide = guide_sheets.get(name, {})
        apply_guide_to_sheet(ws, rows, sheet_guide)
        print(f"  {name}: {len(rows)} rows, {len(sheet_guide.get('named_styles', {}))} named styles")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"\nSaved: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Apply style guide (styles.json) to data.md → styled .xlsx"
    )
    parser.add_argument("data",  help="Input _data.md file")
    parser.add_argument("guide", help="Input _styles.json style guide")
    parser.add_argument("--out", help="Output .xlsx path (default: auto)")
    args = parser.parse_args()

    data_path  = Path(args.data)
    guide_path = Path(args.guide)

    for p in (data_path, guide_path):
        if not p.exists():
            print(f"ERROR: file not found: {p}", file=sys.stderr)
            sys.exit(1)

    if args.out:
        out = Path(args.out)
    else:
        stem = re.sub(r"_data$", "", data_path.stem)
        out = data_path.with_name(stem + "_from_guide.xlsx")

    print(f"Applying {guide_path.name} → {data_path.name} → {out.name} ...")
    convert(data_path, guide_path, out)


if __name__ == "__main__":
    main()
